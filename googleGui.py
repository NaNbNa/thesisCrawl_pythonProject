# -*- codeing = utf-8 -*-
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配`
import requests
#import sqlite3  # 进行SQLite数据库操作
import os  # 判断文件存在
from time import sleep  # 等待间隔
from lxml import etree
import xlsxwriter as xw
from openpyxl import load_workbook, Workbook
from enum import Enum
import tkinter as tk
from tkinter import *
from tkinter import ttk, Tk
import tkinter
import time  
from tkinter import messagebox  # 打开tkiner的消息提醒框
from tkinter import filedialog, simpledialog  # 在Gui中打开文件浏览
import os

class WriteMode(Enum):  
    APPEND = 0  # 追加
    OVERWRITE = 1   #重新写入

def main(*args): # base_url search_word book_path article_num(int)    
    
    base_url = args[0]  #要爬取的网页链接
    # 1.爬取网页
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'

    }
    search_word = args[1]
    mode = None
    if args[2].endswith('.xlsx'):
        book_path = args[2]
        mode = WriteMode.APPEND
    else:
        book_path = args[2] + '/' +search_word + '.xlsx'
        mode = WriteMode.OVERWRITE
    page_num = int(args[3]/10 + 1)
    gui = args[4]

    # 调用函数
    print("开始爬取.......")
    gui.add_text("开始爬取.......")

    if gui.stop_flag== True:
        gui.add_text("中止爬取")
        return
    
    page_info=getData(base_url,search_word,page_num,gui)
    
    # write_to_excel(book_path,page_info,sheet_name,WriteMode.APPEND)
    write_to_excel(book_path,page_info,search_word,gui,mode)

    print("保存完成！")
    gui.add_text("保存完成！")
    
    


# 爬取网页的内容,并解析,填入列表
def getData(baseurl,search_word,page_num,gui,one_page_num=10,): # 默认一页10篇文献
    if gui.stop_flag== True:
        gui.add_text("中止爬取")
        return
    
    datalist = [] 
    for i in range(1, page_num+1):  # 调用多个页面
        if gui.stop_flag== True:
            gui.add_text("中止爬取")
            return
        
        url = baseurl + str(i * one_page_num) + "&q=" + str(search_word) 
        gui.add_text("开始爬取并解析 第" + str(i) + "页的文献......")
        print("开始爬取并解析 第" + str(i) + "页的文献......")
        html = askURL(url)  
        # 2.逐一解析数据
        soup = BeautifulSoup(html, "html.parser")
        items = soup.find_all('div', class_="gs_r gs_or gs_scl")

        number = 0
        for item in items:  
            data = []  # 保存一个文献所有信息

            title = None
            author = None
            type = None
            year = None
            abstract = None
            refer = None
            link = None

            if item.find('h3', class_='gs_rt') is not None:
                title = item.find('h3', class_='gs_rt').text.strip()

            if item.find('div',class_='gs_a') is not None:
                author_and_type = item.find('div',class_='gs_a').text.strip().split('-')
                if author_and_type is not None:
                    author = author_and_type[0].strip()
                if author_and_type is not None:
                    type =  author_and_type[1].split(',')[0]
                if re.search(r'\d{4}', author_and_type[1]) is not None:
                    year = re.search(r'\d{4}', author_and_type[1]).group()   

            if item.find('div', class_='gs_rs') is not None:
                abstract = item.find('div', class_='gs_rs').text.strip()

            if item.find('div', class_='gs_fl gs_flb') is not None:
                if item.find('div', class_='gs_fl gs_flb').find('a', href=lambda href: '/scholar?cites=' in href) is not None:
                    refer = item.find('div', class_='gs_fl gs_flb').find('a', href=lambda href: '/scholar?cites=' in href).text.strip().split()[2]
                    refer = int(refer)
            if item.find('h3', class_='gs_rt') is not None:
                if item.find('h3', class_='gs_rt').find('a') is not None:
                    link = item.find('h3', class_='gs_rt').find('a')['href']
            data.append(title)
            data.append(author)
            data.append(type)
            data.append(year)
            data.append(abstract)
            data.append(refer)
            data.append(link)
            datalist.append(data)

            number = number + 1
            if title is not None:
                print("爬取并解析 文献" + str(number) + ": "+ title +" 成功......")
                gui.add_text("爬取并解析 文献" + str(number) + ": "+ title +" 成功......")
            else: 
                print("爬取并解析 文献" + str(number) + ": " + " 失败......")
                gui.add_text("爬取并解析 文献" + str(number) + ": " + " 失败......")
        print("爬取并解析 第" + str(i) + "页的文献 结束......")
        gui.add_text("爬取并解析 第" + str(i) + "页的文献 结束......")

        if gui.stop_flag== True:
            gui.add_text("中止爬取")
            return
    gui.show_article_list(datalist)
    return datalist


# 得到指定一个URL的网页内容
def askURL(url):
    head = {  # 模拟浏览器头部信息，向豆瓣服务器发送消息
        "User-Agent": "Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36"
    }
    # 用户代理，表示告诉豆瓣服务器，我们是什么类型的机器、浏览器（本质上是告诉浏览器，我们可以接收什么水平的文件内容）

    try:  
        response = requests.get(url, headers=head)  
        response.raise_for_status()  # 如果请求失败则抛出HTTPError异常  
        html = response.text  
    except requests.exceptions.HTTPError as http_err:  
        print(f"HTTP error occurred: {http_err}")  
    except requests.exceptions.RequestException as err:  
        print(f"An error occurred: {err}")  
    else:  
        return html 


def write_to_excel(book_path, info,search_word,gui,mode=WriteMode.OVERWRITE):
    if gui.stop_flag== True:
        gui.add_text("中止爬取")
        return
    
    sheet = None
    if not os.path.exists(book_path):  
        book = Workbook()  
        # 删除默认创建的Sheet（通常名为'Sheet'）  
        book.remove(book.active)  
    else:  
        book = load_workbook(book_path)  

        # 选择--追加,或者重新写入.xlsx
    try:  
        if mode == WriteMode.OVERWRITE:    # 默认重新写入  
            try:  
                book.remove(book[search_word])  # 删除原有的sheet
            except KeyError:  
                # 工作表不存在时忽略错误  
                pass
        sheet = book[search_word] # 新建
    except KeyError:  
        sheet = book.create_sheet(title=search_word)
    
    col = ['标题', '作者', '类型', '出版日期', '摘要', '被引量', '链接']  # 设置表头

    # 设置表头，只在工作表是新创建的时候写入  
    if sheet.max_row == 1 and sheet.max_column == 1:  
        for col_idx, header_name in enumerate(col, start=1):  
            sheet.cell(row=1, column=col_idx, value=header_name) 

    # append,参数是一个列表或者元组
    for data in info:
        if gui.stop_flag== True:
            gui.add_text("中止爬取")
            return  
        sheet.append(data)

    try:  
        book.save(book_path)  
        print("文件保存成功！")
        gui.add_text("文件保存成功！")
    except PermissionError as pe:  
        print(f"保存文件时发生权限错误: {pe}")  
        gui.add_text(f"保存文件时发生权限错误: {pe}")
    except Exception as e:  
        print(f"保存文件时发生错误: {e}")
        gui.add_text(f"保存文件时发生错误: {e}")
    return True


