import requests  # 发送请求
import os  # 判断文件存在
from time import sleep  # 等待间隔
from lxml import etree
from openpyxl import load_workbook, Workbook
from enum import Enum

def main():
    base_url = 'http://search.cnki.com.cn/Search/ListResult'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'

    }
    page_num =2
    search_word = "金融科技"
    book_path = search_word + '.xlsx'
    # 调用函数
    print("开始爬取.......")
    for page in  range(1,page_num+1):

        page_text =get_page_text(base_url,headers,search_word,page)
        page_info =parse_page_text(page_text,headers)
        sheet_name = search_word + "-" + str(page)

        print("开始保存:第" +  str(page) + "页")
        # write_to_excel(book_path,page_info,sheet_name,WriteMode.APPEND)
        write_to_excel(book_path,page_info,sheet_name)

    print("保存完成！")

class WriteMode(Enum):  
    APPEND = 0  # 追加
    OVERWRITE = 1   #重新写入

# url: 需要进行POST请求的URL地址，是一个字符串类型
# headers: HTTP请求头信息，通常是一个字典类型，包含如Content-Type, Authorization等键值对
# search_word: 用于搜索的关键词，是一个字符串类型
# page_num: 需要获取的页面编号，是一个整数类型
# page_text: 返回的页面文本内容，是一个字符串类型

# get_page_text 函数是一个用于向指定URL发送POST请求以获取页面文本内容的函数
def get_page_text(url, headers, search_word, page_num):
    # 定义一个字典data，用于存放POST请求的参数  
    data = {
        'searchType': 'MulityTermsSearch',
        'ArticleType': '',
        'ReSearch': '',
        'ParamIsNullOrEmpty': 'false',
        'Islegal': 'false',
        'Content': '',
        'Theme': search_word,
        'Title': '',
        'KeyWd': '',
        'Author': '',
        'SearchFund': '',
        'Originate': '',
        'Summary': '',
        'PublishTimeBegin': '',
        'PublishTimeEnd': '',
        'MapNumber': '',
        'Name': '',
        'Issn': '',
        'Cn': '',
        'Unit': '',
        'Public': '',
        'Boss': '',
        'FirstBoss': '',
        'Catalog': '',
        'Reference': '',
        'Speciality': '',
        'Type': '',
        'Subject': '',
        'SpecialityCode': '',
        'UnitCode': '',
        'Year': '',
        'AcefuthorFilter': '',
        'BossCode': '',
        'Fund': '',
        'Level': '',
        'Elite': '',
        'Organization': '',
        'Order': '1',
        'Page': str(page_num),
        'PageIndex': '',
        'ExcludeField': '',
        'ZtCode': '',
        'Smarts': '',
    }

    # 使用requests库的post方法发送POST请求，传入url, headers和data作为参数 
    response = requests.post(url=url, headers=headers, data=data)
    # 获取响应的文本内容  
    page_text = response.text
    return page_text


# list 转 str
def list_to_str(my_list):  
    # 使用列表推导式将嵌套列表展平  
    flattened_list = [item for sublist in my_list for item in sublist if isinstance(item, str)]  
    # 确保展平后的列表中只包含字符串  
    if all(isinstance(item, str) for item in flattened_list):  
        my_str = "".join(flattened_list)  
        return my_str  
    else:  
        raise ValueError("List contains non-string items after flattening.")


# url: 一个字符串，表示要从中提取摘要的网页的URL
# abstract: 一个列表，包含从网页中提取的摘要文本。
# 列表中的每个元素都是一个字符串，代表从class为"xx_font"的<div>元素中捕获的一段文本内容

# 从给定的网页URL中提取摘要文本,在parse_page_text调用
def get_abstract(url,headers):  
    try:  
        # 发送GET请求  
        response = requests.get(url, headers=headers)  
        # 检查请求是否成功  
        response.raise_for_status()  
        # 提取网页文本内容  
        page_text = response.text  
        # 尝试解析网页内容  
        try:  
            tree = etree.HTML(page_text)  
        except etree.ParseError:  
            # 处理HTML解析错误  
            return [], "Failed to parse the HTML content." 
        # 使用XPath查询提取摘要  
        try:  
            abstract = tree.xpath('//div[@class="xx_font"]//text()')  
        except etree.XPathEvalError:  
            # 处理XPath查询错误  
            return [], "Invalid XPath query or no matching elements found."  
        # 返回摘要文本列表  
        return abstract, "Success"  
        
    except requests.exceptions.RequestException as e:  
        # 处理网络请求错误  
        return [], f"Network request failed: {e}"    
    except Exception as e:  
        # 处理其他未知异常  
        return [], f"An unexpected error occurred: {e}"  

# xpath解析,接受html,返回[]
# page_text: HTML格式的字符串，通常是从网页上抓取下来的
# page_info: 这是一个二维列表，
# 其中每个内部列表都包含一条从HTML中解析出来的信息
# 包括标题、作者、文献来源、文献类型、出版日期、摘要、关键词、下载量、被引量以及链接
# HTML解析:
# tree = etree.HTML(page_text): 
# 使用etree.HTML解析HTML文本，生成一个可以进行XPath查询的对象
# XPath查询:
# item_list = tree.xpath('//div[@class="list-item"]'): 
# 查询所有class为"list-item"的<div>元素
# 遍历和提取信息
    
# parse_page_text 函数的功能是从给定的 HTML 页面文本（page_text）中解析出特定结构的信息，
# 并将这些信息整理成一个列表的列表（page_info）返回
def parse_page_text(page_text,headers):
    tree = etree.HTML(page_text)
    item_list = tree.xpath('//div[@class="list-item"]')
    page_info = []
    print("解析一页信息开始......")
    number = 0      # 文献序号
    for item in item_list:
        # 标题
        title = list_to_str(item.xpath(
            './p[@class="tit clearfix"]/a[@class="left"]/@title'))
        # 链接
        link = 'https:' +\
            list_to_str(item.xpath(
                './p[@class="tit clearfix"]/a[@class="left"]/@href'))
        # 作者
        author = list_to_str(item.xpath(
            './p[@class="source"]/span[1]/@title'))
        # 出版日期
        date = list_to_str(item.xpath(
            './p[@class="source"]/span[last()-1]/text() | ./p[@class="source"]/a[2]/span[1]/text() '))
        # 关键词
        keywords = list_to_str(item.xpath(
            './div[@class="info"]/p[@class="info_left left"]/a[1]/@data-key'))
        # 摘要
        abstract = list_to_str(get_abstract(url=link,headers=headers))
        # 文献来源
        paper_source = list_to_str(item.xpath(
            './p[@class="source"]/span[last()-2]/text() | ./p[@class="source"]/a[1]/span[1]/text() '))
        # 文献类型
        paper_type = list_to_str(item.xpath(
            './p[@class="source"]/span[last()]/text()'))
        # 下载量
        download = list_to_str(item.xpath(
            './div[@class="info"]/p[@class="info_right right"]/span[@class="time1"]/text()'))
        # 被引量
        refer = list_to_str(item.xpath(
            './div[@class="info"]/p[@class="info_right right"]/span[@class="time2"]/text()'))

        item_info = [i.strip() for i in [title, author, paper_source, paper_type, date, abstract, keywords, download, refer, link]]
        page_info.append(item_info)

        number = number + 1
        print("解析 文献" + str(number) + ": "+ title +" 成功......")
    print("解析一页信息完成......")
    return page_info

# 保存数据
# workbook：一个已存在的Excel工作簿对象。
# info：一个包含多个数据行的列表，其中每一行都是一个包含标题对应信息的列表。
# search_word：一个字符串，用于作为新创建的工作表的名称。
def write_to_excel(book_path, info,sheet_name,mode=WriteMode.OVERWRITE):
    sheet = None
    if not os.path.exists(book_path):  
        book = Workbook()  
        # 删除默认创建的Sheet（通常名为'Sheet'）  
        book.remove(book.active)  
    else:  
        book = load_workbook(book_path)  

        # 选择--追加,或者重新写入.xlsx
    try:  
        if mode == WriteMode.OVERWRITE:    # 默认重新写入  
            try:  
                book.remove(book[sheet_name])  # 删除原有的sheet
            except KeyError:  
                # 工作表不存在时忽略错误  
                pass
        sheet = book[sheet_name] # 新建
    except KeyError:  
        sheet = book.create_sheet(title=sheet_name)

    # col = ['title', 'author', 'paper_source', 'paper_type', 'date', 'abstract', 'keywords', 'download', 'refer', 'link']  # 设置表头
    col = ['标题', '作者', '文献来源', '类型', '出版日期', '摘要', '关键词', '下载量', '被引量', '链接']  # 设置表头

    # 设置表头，只在工作表是新创建的时候写入  
    if sheet.max_row == 1 and sheet.max_column == 1:  
        for col_idx, header_name in enumerate(col, start=1):  
            sheet.cell(row=1, column=col_idx, value=header_name) 
    
    # append,参数是一个列表或者元组
    for data in info:  
        sheet.append(data)

    try:  
        book.save(book_path)  
        print("文件保存成功！")  
    except PermissionError as pe:  
        print(f"保存文件时发生权限错误: {pe}")  
    except Exception as e:  
        print(f"保存文件时发生错误: {e}")
    return True


if __name__ == "__main__":  # 当程序执行时
    # 基础配置,模拟url,headers
    main()
    print("爬取完毕！")
